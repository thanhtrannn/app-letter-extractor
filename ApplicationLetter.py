# Script uses selenium and chrome web driver to log into DRUPAL with provided credentials,
# then extract students MMFAP letters and rename them with appropriate standard. List of URL
# and student information is taken from Calculated_Table extraction
# Thanh Tran - June 3, 2019

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import selenium.webdriver.support.ui as ui
from selenium.webdriver.chrome.options import Options
from selenium.webdriver import Chrome
import os
import time
import glob
import urllib 
from openpyxl import load_workbook
import sys
from tkinter import messagebox
from tkinter import filedialog
from tkinter import *

dataDict = {}
pathSelected = False

# GUI Section

# used to select folder for letters to be extracted in
def browse_button():
    path = filedialog.askdirectory(title='Select a folder')
    if path is not None:
        global pathSelected
        folderString.set(path.replace("/", "\\"))
        pathSelected = True
    
# used to select calculated file to be used in script
def filebrowse_button():
    global file
    file = filedialog.askopenfile(parent=root, mode='rb', title='Select a file', filetypes=[('Excel file', '*.xlsx;*.xlsm;*xls')])
    if file is not None:
        fileString.set(file.name)
    
root = Tk()
root.title('MMFAP Application Letter Extraction')

# set Labels
Label(root, text='MMFAP Application Letter Extraction', font=('Aria', 12, 'bold')).grid(row=0, padx=10, pady=20, columnspan=3)
Label(root, text='DRUPAL login URL: ').grid(row=1, padx=10, pady=10, sticky="e")
Label(root, text='DRUPAL username: ').grid(row=2, padx=10, pady=10, sticky="e")
Label(root, text='DRUPAL password: ').grid(row=3,padx=10, pady=10, sticky="e")
Label(root, text='Extract Letters To: ').grid(row=4, padx=10, pady=10, sticky="e")
Label(root, text='Calculated File: ').grid(row=5, padx=10, pady=10, sticky="e")
Label(root, text='Seconds between downloads: ').grid(row=6, padx=10, pady=5, sticky="e")
# set Buttons
browseButton = Button(root, text='Browse', command=browse_button, width = 15).grid(row=4, column=2, pady= 5, padx=10)
browseFileButton = Button(root, text='Browse', command=filebrowse_button, width = 15).grid(row=5, column=2)
startButton = Button(root, text='Start', command=root.quit, width = 15, bg="medium sea green", fg="white").grid(row=7, column=1, pady=20, padx=10)
# set Variables
drupalURLString = StringVar()
drupalURLString.set('https://www.mohawkcollege.ca/user/login')
folderString = StringVar()
fileString = StringVar()
timeDLString = IntVar()
timeDLString.set(3)
username = StringVar()
password = StringVar()
# set Fields
Entry(root, textvariable=drupalURLString, width=50).grid(row=1, column=1)
Entry(root, width=50, textvariable=username).grid(row=2, column=1)
Entry(root, width=50, textvariable=password).grid(row=3, column=1)
Entry(root, width=50, state='readonly', textvariable=folderString).grid(row=4, column=1)
Entry(root, width=50, state='readonly', textvariable=fileString).grid(row=5, column=1)
Entry(root, textvariable=timeDLString, width=25).grid(row=6, column=1, sticky='w')

root.mainloop()

# make directory to hold extract
# path = os.getcwd() + '\\' + folder
# if not os.path.exists(path):
#     os.makedirs(path)
    
if getattr(sys, 'frozen', False) :
    # running in a bundle
    base_dir = sys._MEIPASS
else:
    # running normally
    base_dir = os.path.dirname(os.path.abspath(__file__))

chromedriver_path = os.path.join(base_dir, 'chromedriver')
# create folder in current directory to hold letters if folder was not selected
if pathSelected is False:
    folderString.set(os.getcwd() + '\\Application Letters')
    if not os.path.exists(folderString.get()):
        os.makedirs(folderString.get())

# open calculated file
try:
    if file is not None:
        wb_obj = load_workbook(filename = file)
        wsheet = wb_obj['Calculated_Table']
        # cycle through rows
        for key, *values in wsheet.iter_rows():
            dataDict[key.value] = [v.value for v in values]
except:
    messagebox.showinfo("Error", "Calculated File, NOT FOUND")
    exit()

# chrome driver settings
chrome_options = webdriver.ChromeOptions()
prefs = {'download.default_directory' : folderString.get()}
chrome_options.add_experimental_option('prefs', prefs)
driver = webdriver.Chrome(chromedriver_path,chrome_options=chrome_options)

# get header columns
count = 0
firstNameCol = 0
lastNameCol = 0
studentNumCol = 0
appLetterCol = 0
for heading in dataDict[next(iter(dataDict))]:
    if heading == 'Banner_First_Name':
        firstNameCol = count
    if heading == 'Banner_Last_Name':
        lastNameCol = count
    if heading == 'CorrectedMohawkCollegeStudentID':
        studentNumCol = count
    if heading == 'ApplicationLetter':
        appLetterCol = count
    count += 1

# begin initial login of DRUPAL
driver.get(drupalURLString.get())
try:
    ui.WebDriverWait(driver, 10).until(lambda x: x.find_element_by_id('edit-name'))
    driver.find_element_by_id("edit-name").click()
    driver.find_element_by_id("edit-name").clear()
    driver.find_element_by_id("edit-name").send_keys(username.get())
    driver.find_element_by_id("edit-pass").click()
    driver.find_element_by_id("edit-pass").clear()
    driver.find_element_by_id("edit-pass").send_keys(password.get())
    driver.find_element_by_id("edit-submit").click()
    # wait for web page to load after logging - ensure welcome screen is presented
    driver.find_element_by_xpath("//h1[@class='page-title' and text()='" + username.get().lower() + "']")
except:
    driver.close()
    driver.quit()
    messagebox.showinfo("Error", "Login is not accepted")
    exit()

# cycle through records
for x in dataDict:
    # ignore empty cells
    if dataDict[x][appLetterCol] is not None and 'https' in dataDict[x][appLetterCol]:
        driver.get(dataDict[x][appLetterCol])
        # add timer to allow for download to complet
        time.sleep(int(timeDLString.get()))
        # get exacted downloaded file name
        file_name = dataDict[x][appLetterCol]
        file_name = file_name.split("/")
        file_name = urllib.parse.unquote(file_name[len(file_name) - 1])
        file_to_rename = os.path.join(folderString.get(), dataDict[x][firstNameCol] + "." + dataDict[x][lastNameCol] + "." + str(dataDict[x][studentNumCol]) +".pdf")
        #search for the download and rename
        for filename in os.listdir(folderString.get()):
            # check for file renaming, if file exist then rename with appending number
            if filename.startswith(file_name) and not os.path.exists(file_to_rename):
                    os.rename(os.path.join(folderString.get(), filename), file_to_rename)
            elif filename.startswith(file_name) and os.path.exists(file_to_rename):
                trigger = True
                count = 1
                # continue searching for duplicates
                while trigger:
                    if not os.path.exists(os.path.join(folderString.get(), dataDict[x][firstNameCol] + "." + dataDict[x][lastNameCol] + "." + str(dataDict[x][studentNumCol]) + "_" + str(count) + ".pdf")):
                        os.rename(os.path.join(folderString.get(), filename), os.path.join(folderString.get(), dataDict[x][firstNameCol] + "." + dataDict[x][lastNameCol] + "." + str(dataDict[x][studentNumCol]) + "_" + str(count) + ".pdf"))
                        trigger = False
                    count = count + 1
            else:
                continue
        
messagebox.showinfo("Extraction have been completed, check extraction folder to ensure letters are there.")
driver.close()
driver.quit()


